import os
import io
import traceback
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import range_boundaries
from django.http import JsonResponse

from django.http import HttpResponse, JsonResponse
import pandas as pd
import io
import traceback
import os

def export_excel(
    data_summary,
    data_detail,
    filename: str = "Slow_moving.xlsx",
    store_: str = "downClient",
    path_: str = None
):
    """
    - store_ == "downClient": tải file về client (HttpResponse)
    - store_ != "downClient": lưu file lên server tại path_ và trả JsonResponse thông tin
    """

    try:
        # 1) Tạo DataFrame từ dữ liệu (tránh lỗi nếu None)
        df1 = pd.DataFrame(data_summary or [])
        df2 = pd.DataFrame(data_detail or [])

        # 2) Ghi workbook vào buffer in-memory
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            wrote_any = False  # <<< THÊM DÒNG NÀY

            # Summary
            if df1 is not None and (not df1.empty or len(df1.columns) > 0):  # ← sửa điều kiện này
                df1.to_excel(writer, index=False, sheet_name='Summary', startrow=1)

                # Các header gộp/định dạng (giữ nguyên theo code gốc)
                fill_range(writer, sheet_name="Summary", cell_range="A2:J2", fill_color="FFFF99")
                format_header(writer, sheet_name="Summary", cell_range="K1:M1",
                              text="Stock End\n Month N-1", fill_color="FFD700", border=True)
                fill_range(writer, sheet_name="Summary", cell_range="K2:M2", fill_color="FFD700")
                format_header(writer, sheet_name="Summary", cell_range="N1:P1",
                              text="Normal stock\n 1-3 month", fill_color="FBE2D5", border=True)         
                format_header(writer, sheet_name="Summary", cell_range="Q1:S1",
                              text="Excess stock\n 4-6 month", fill_color="FBE2D5", border=True) 
                fill_range(writer, sheet_name="Summary", cell_range="N2:S2", fill_color="FBE2D5")
                format_header(writer, sheet_name="Summary", cell_range="T1:V1",
                              text="Non moving\n (stock at least 6month\n + no use in 6 months)",
                              fill_color="FBE2D5", font_color="FF0000", border=True)
                fill_range(writer, sheet_name="Summary", cell_range="T2:V2", fill_color="FBE2D5", font_color="FF0000")
                format_header(writer, sheet_name="Summary", cell_range="W1:Y1",
                              text="Slow moving\n stock 7-12 month", fill_color="FBE2D5", border=True)
                format_header(writer, sheet_name="Summary", cell_range="Z1:AB1",
                              text="Slow moving\n stock 13-24 month", fill_color="FBE2D5", border=True)
                format_header(writer, sheet_name="Summary", cell_range="AC1:AE1",
                              text="Slow moving\n stock Over 25-36 month", fill_color="FBE2D5", border=True)  
                format_header(writer, sheet_name="Summary", cell_range="AF1:AH1",
                              text="Slow moving\n stock Over 36 month", fill_color="FBE2D5", border=True)
                
                fill_range(writer, sheet_name="Summary", cell_range="W2:AH2", fill_color="FBE2D5")
                format_header(writer, sheet_name="Summary", cell_range="AI1:AK1",
                              text="Total Slow\n moving", fill_color="FBE2D5",
                              font_color="FF0000", border=True)
                format_header(writer, sheet_name="Summary", cell_range="AL1:AN1",
                              text="Total Slow\n + Non moving", fill_color="FBE2D5",
                              font_color="FF0000", border=True)
                fill_range(writer, sheet_name="Summary", cell_range="AI2:AN2", fill_color="FBE2D5", font_color="FF0000")

            # Detail
            if df2 is not None and (not df2.empty or len(df2.columns) > 0):  # ← sửa điều kiện này
                df2.to_excel(writer, index=False, sheet_name='Detail', startrow=1)
                # Tô nền vàng cho vùng A1:V1
                fill_range(writer, sheet_name="Detail", cell_range="A2:Y2", fill_color="FFFF99")
                format_header(writer, sheet_name="Detail", cell_range="Z1:AB1",
                              text="Stock End\n Month N-1", fill_color="FFD700", border=True)
                fill_range(writer, sheet_name="Detail", cell_range="Z2:AB2", fill_color="FFD700")
                format_header(writer, sheet_name="Detail", cell_range="AC1:AD1",
                              text="Month N-1", fill_color="C1F0C8", border=True)
                format_header(writer, sheet_name="Detail", cell_range="AE1:AF1",
                              text="Month N-1\n to N-3", fill_color="C1F0C8", border=True)
                format_header(writer, sheet_name="Detail", cell_range="AG1:AH1",
                              text="Month N-4\n to N-6", fill_color="C1F0C8", border=True)
                format_header(writer, sheet_name="Detail", cell_range="AI1:AJ1",
                              text="Month N-7\n to N-12", fill_color="C1F0C8", border=True)
                format_header(writer, sheet_name="Detail", cell_range="AK1:AL1",
                              text="Month N-13\n to N-24", fill_color="C1F0C8", border=True)
                format_header(writer, sheet_name="Detail", cell_range="AM1:AN1",
                              text="Month N-25\n to N-36", fill_color="C1F0C8", border=True)
                fill_range(writer, sheet_name="Detail", cell_range="AC2:AN2", fill_color="C1F0C8")
                format_header(writer, sheet_name="Detail", cell_range="AO1:AO2",
                              text="BF result\n tháng N-1", fill_color="B1A0C7", border=True)
                format_header(writer, sheet_name="Detail", cell_range="AP1:AP2",
                              text="BF result\n tháng N-1 đến N-6", fill_color="B1A0C7", border=True)
                format_header(writer, sheet_name="Detail", cell_range="AQ1:AS1",
                              text="Normal stock\n 1-3 month", fill_color="FBE2D5", border=True)
                format_header(writer, sheet_name="Detail", cell_range="AT1:AV1",
                              text="Excess stock\n 4-6 month", fill_color="FBE2D5", border=True)
                format_header(writer, sheet_name="Detail", cell_range="AW1:AY1",
                              text="Non moving\n (stock at least 6month\n + no use in 6 months)",
                              fill_color="FBE2D5", font_color="FF0000", border=True)    
                format_header(writer, sheet_name="Detail", cell_range="AZ1:BB1",
                              text="Slow moving\n stock 7-12 month", fill_color="FBE2D5", border=True)
                format_header(writer, sheet_name="Detail", cell_range="BC1:BE1",
                              text="Slow moving\n stock 13-24 month", fill_color="FBE2D5", border=True)
                format_header(writer, sheet_name="Detail", cell_range="BF1:BH1",
                              text="Slow moving\n stock Over 25-36 month", fill_color="FBE2D5", border=True)
                format_header(writer, sheet_name="Detail", cell_range="BI1:BK1",
                              text="Slow moving\n stock Over 36 month", fill_color="FBE2D5", border=True)
                format_header(writer, sheet_name="Detail", cell_range="BL1:BN1",
                              text="Total Slow\n moving", fill_color="FBE2D5",
                              font_color="FF0000", border=True)
                format_header(writer, sheet_name="Detail", cell_range="BO1:BQ1",
                              text="Total Slow\n + Non moving", fill_color="FBE2D5",
                              font_color="FF0000", border=True)     
                format_header(writer, sheet_name="Detail", cell_range="BR1:BS1",
                              text="Other plants using", fill_color="CAEDFB", border=True)
                fill_range(writer, sheet_name="Detail", cell_range="BR2:BS2", fill_color="CAEDFB")
                format_header(writer, sheet_name="Detail", cell_range="BT1:BU1",
                              text="EOL", fill_color="FFD580", border=True)
                format_header(writer, sheet_name="Detail", cell_range="BV1:BW1",
                              text="Discont", fill_color="FFD580", border=True)
                fill_range(writer, sheet_name="Detail", cell_range="AQ2:AV2", fill_color="FBE2D5")
                fill_range(writer, sheet_name="Detail", cell_range="AW2:AY2", fill_color="FBE2D5", font_color="FF0000")
                fill_range(writer, sheet_name="Detail", cell_range="AZ2:BK2", fill_color="FBE2D5")
                fill_range(writer, sheet_name="Detail", cell_range="BL2:BQ2", fill_color="FBE2D5", font_color="FF0000")
                fill_range(writer, sheet_name="Detail", cell_range="BT2:BW2", fill_color="FFD580")



        # Nếu không có sheet nào được ghi (khi detail rỗng, summary cũng trống)
        if not wrote_any:  # ← thêm đoạn fallback này
            pd.DataFrame({'Thông báo': ['Không có dữ liệu để xuất']}).to_excel(
                writer, index=False, sheet_name='Report'
            )


        # 3) Đọc dữ liệu nhị phân từ buffer
        buffer.seek(0)  # Quan trọng
        excel_data = buffer.read()
        buffer.close()

        # 4) Điều kiện xuất/ lưu
        if store_ == "downClient":
            # Tải về client
            response = HttpResponse(
                excel_data,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            response["Access-Control-Expose-Headers"] = "Content-Disposition"
            return response
        else:
            # Lưu lên server tại path_
            if not path_:
                return JsonResponse(
                    {"detail": "Thiếu tham số path_ khi lưu file lên server."},
                    status=400,
                )

            # Tạo thư mục nếu chưa có
            os.makedirs(path_, exist_ok=True)

            filename2 = f"{filename}.xlsx"               # Nối thêm đuôi .xlsx vì ở view k để đuôi .xlsx

            # Ghép đường dẫn đầy đủ
            filepath = os.path.join(path_, filename2)

            # Ghi file ra đĩa
            with open(filepath, "wb") as f:
                f.write(excel_data)
            
            # Trả thông tin lưu thành công
            return JsonResponse(
                {
                    "detail": "Lưu file thành công trên server.",
                    "filepath": filepath,
                    "filename": filename2,
                },
                status=200,
            )

    except Exception as e:
        # Log server
        print(f"Lỗi khi xuất Excel: {e}")
        traceback.print_exc()

        # Phản hồi lỗi cho client
        return JsonResponse(
            {"detail": f"Lỗi khi xuất file Excel: {str(e)}"},
            status=500,
        )
def format_header(writer, sheet_name, cell_range, text, font_size=11, fill_color=None, font_color="000000", border=False):
    """
    Format header cho một sheet Excel.
    
    Args:
        writer: ExcelWriter object (pandas.ExcelWriter).
        sheet_name (str): Tên sheet muốn chỉnh sửa.
        cell_range (str): Khoảng ô cần merge, ví dụ 'W1:Y1'.
        text (str): Nội dung muốn hiển thị.
        font_size (int, optional): Kích thước font, mặc định = 11.
        fill_color (str, optional): Màu nền hex (VD: 'FFFF00' = vàng).
        font_color (str, optional): Màu chữ hex (VD: 'FF0000' = đỏ).
        border (bool, optional): Nếu True sẽ vẽ viền quanh ô/merge.
    """
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    worksheet.merge_cells(cell_range)
    start_cell = cell_range.split(":")[0]  # lấy ô bắt đầu (VD: 'W1')

    worksheet[start_cell] = text
    worksheet[start_cell].font = Font(bold=True, size=font_size, color=font_color)
    worksheet[start_cell].alignment = Alignment(horizontal='center', vertical='center')
    # Nếu có màu nền thì bôi màu
    if fill_color:
        worksheet[start_cell].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    
    # Nếu bật border thì thêm viền quanh vùng merge
    if border:
        thin = Side(border_style="thin", color="000000")
        border_style = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Áp border cho tất cả ô trong vùng merge
        min_col = worksheet[cell_range.split(":")[0]].column
        min_row = worksheet[cell_range.split(":")[0]].row
        max_col = worksheet[cell_range.split(":")[1]].column
        max_row = worksheet[cell_range.split(":")[1]].row

        for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, 
                                       min_col=min_col, max_col=max_col):
            for cell in row:
                cell.border = border_style



def _normalize_hex(color: str) -> str:
    """
    Chuẩn hóa mã màu hex cho openpyxl.
    - Nếu truyền 'RRGGBB' sẽ tự thêm alpha '00' -> '00RRGGBB'
    - Nếu đã là 'AARRGGBB' thì giữ nguyên.
    """
    color = color.strip().lstrip("#")
    if len(color) == 6:
        return "00" + color.upper()
    if len(color) == 8:
        return color.upper()
    raise ValueError("Mã màu phải ở dạng RRGGBB hoặc AARRGGBB")

def fill_range(writer, sheet_name: str, cell_range: str, fill_color: str, font_color: str = None):
    """
    Tô màu toàn bộ vùng ô (cell_range) trên sheet (sheet_name).

    Args:
        writer: pandas.ExcelWriter (engine openpyxl).
        sheet_name (str): Tên sheet.
        cell_range (str): Vùng ô, ví dụ 'B2:F10' hoặc 'A1:A1'.
        fill_color (str): Màu nền (hex), ví dụ 'FFFF00' (vàng) hoặc '80FFFF00' (vàng nhạt có alpha).
        font_color (str, optional): Nếu muốn, đặt màu chữ cho cả vùng (hex).
    """
    wb = writer.book
    ws = writer.sheets[sheet_name]

    # Chuẩn hóa màu và tạo PatternFill
    fill_hex = _normalize_hex(fill_color)
    fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")

    # Lặp qua toàn bộ ô trong range và áp màu
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.fill = fill
            if font_color:
                cell.font = Font(
                    name=(cell.font.name if cell.font else None),
                    size=(cell.font.size if cell.font else None),
                    bold=(cell.font.bold if cell.font else None),
                    italic=(cell.font.italic if cell.font else None),
                    color=_normalize_hex(font_color)
                )






def export_excel_FG(
    data,
    filename: str = "Slow_moving.xlsx",
    store_: str = "downClient",
    path_: str = None
):
    """
    - store_ == "downClient": tải file về client (HttpResponse)
    - store_ != "downClient": lưu file lên server tại path_ và trả JsonResponse thông tin
    """

    try:
        # 1) Tạo DataFrame từ dữ liệu (tránh lỗi nếu None)
        df1 = pd.DataFrame(data or [])

        # 2) Ghi workbook vào buffer in-memory
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            wrote_any = False  # <<< THÊM DÒNG NÀY

            # Data
            if df1 is not None and (not df1.empty or len(df1.columns) > 0):  # ← sửa điều kiện này
                df1.to_excel(writer, index=False, sheet_name='Data', startrow=1)

                # Các header gộp/định dạng (giữ nguyên theo code gốc)
                fill_range(writer, sheet_name="Data", cell_range="A2:E2", fill_color="FFFF99")
                fill_range(writer, sheet_name="Data", cell_range="F2:F2", fill_color="E26B0A")
                fill_range(writer, sheet_name="Data", cell_range="G2:L2", fill_color="B7DEE8")
                fill_range(writer, sheet_name="Data", cell_range="M2:M2", fill_color="31869B")
                fill_range(writer, sheet_name="Data", cell_range="N2:O2", fill_color="C4D79B")
                fill_range(writer, sheet_name="Data", cell_range="P2:P2", fill_color="E26B0A")
                fill_range(writer, sheet_name="Data", cell_range="Q2:Q2", fill_color="FFFF00")

        # Nếu không có sheet nào được ghi
        if not wrote_any:  # ← thêm đoạn fallback này
            pd.DataFrame({'Thông báo': ['Không có dữ liệu để xuất']}).to_excel(
                writer, index=False, sheet_name='Report'
            )


        # 3) Đọc dữ liệu nhị phân từ buffer
        buffer.seek(0)  # Quan trọng
        excel_data = buffer.read()
        buffer.close()

        # 4) Điều kiện xuất/ lưu
        if store_ == "downClient":
            # Tải về client
            response = HttpResponse(
                excel_data,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            response["Access-Control-Expose-Headers"] = "Content-Disposition"
            return response
        else:
            # Lưu lên server tại path_
            if not path_:
                return JsonResponse(
                    {"detail": "Thiếu tham số path_ khi lưu file lên server."},
                    status=400,
                )

            # Tạo thư mục nếu chưa có
            os.makedirs(path_, exist_ok=True)

            filename2 = f"{filename}.xlsx"               # Nối thêm đuôi .xlsx vì ở view k để đuôi .xlsx

            # Ghép đường dẫn đầy đủ
            filepath = os.path.join(path_, filename2)

            # Ghi file ra đĩa
            with open(filepath, "wb") as f:
                f.write(excel_data)
            
            # Trả thông tin lưu thành công
            return JsonResponse(
                {
                    "detail": "Lưu file thành công trên server.",
                    "filepath": filepath,
                    "filename": filename2,
                },
                status=200,
            )

    except Exception as e:
        # Log server
        print(f"Lỗi khi xuất Excel: {e}")
        traceback.print_exc()

        # Phản hồi lỗi cho client
        return JsonResponse(
            {"detail": f"Lỗi khi xuất file Excel: {str(e)}"},
            status=500,
        )